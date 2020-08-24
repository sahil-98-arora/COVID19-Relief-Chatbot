import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect('bms.db')
wb = load_workbook('bms_menu.xlsx')
ws = wb['items']

conn.execute("create table if not exists bms1 (gender text, choice text, price int)")

for i in range(1,17):
    temp_str = "insert into bms1 (gender, choice, price) values ('{0}', '{1}', '{2}')".format(ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value)
    conn.execute(temp_str)

conn.commit()

content = conn.execute("select * from bms1")
for i in content:
    print(i)


#conn = sqlite3.connect('bms.db')
#user_message = "for men"

#print("User message : ", user_message)
#if "men" in user_message:
#    exe_str = "Select choice, price from bms1 where gender is '{0}'".format('men')
#elif 'women' in user_message:
#    exe_str = "Select choice, price from bms1 where gender is '{0}'".format('women')

#content = conn.execute(exe_str)
#content_text = ''
#for index, value in enumerate(content):
#    content_text += str(index + 1) + ") " + str(value[0]) + "  ----  " + str(value[1]) + "/-\n"
#print(content_text)