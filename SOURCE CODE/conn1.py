import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect('nm.db')
wb = load_workbook('nm_menu.xlsx')
ws = wb['items']

conn.execute("create table if not exists nm1 (category text, choice text, price int)")

for i in range(1,18):
    temp_str = "insert into nm1 (category, choice, price) values ('{0}', '{1}', '{2}')".format(ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value)
    conn.execute(temp_str)

conn.commit()

content = conn.execute("select * from nm1")
for i in content:
    print(i)

